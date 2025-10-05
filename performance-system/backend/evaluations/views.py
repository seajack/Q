from django.shortcuts import render
from rest_framework import viewsets, status
from rest_framework.decorators import api_view, action
from rest_framework.response import Response
from django.db.models import Count, Q, Avg, Sum, F
from django.utils import timezone
from datetime import datetime, timedelta
from .models import (
    EvaluationCycle, EvaluationIndicator, EvaluationTask,
    EvaluationScore, EvaluationResult, Employee, EvaluationRule, ManualEvaluationAssignment,
    PositionWeight
)
from .serializers import (
    EvaluationCycleSerializer, EvaluationIndicatorSerializer,
    EvaluationTaskSerializer, EvaluationScoreSerializer,
    EvaluationResultSerializer, EvaluationStatsSerializer, EmployeeSerializer,
    EvaluationRuleSerializer, ManualEvaluationAssignmentSerializer,
    PositionWeightSerializer
)
import requests
from django.conf import settings
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter


class EvaluationCycleViewSet(viewsets.ModelViewSet):
    """考核周期管理"""
    queryset = EvaluationCycle.objects.all().order_by('-created_at')
    serializer_class = EvaluationCycleSerializer
    # 服务端筛选/检索/排序
    filterset_fields = ['status']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'start_date', 'end_date', 'name']
    
    def perform_create(self, serializer):
        """创建考核周期时自动设置创建者"""
        # 暂时使用第一个超级用户作为创建者，实际应该使用当前登录用户
        from django.contrib.auth.models import User
        creator = User.objects.filter(is_superuser=True).first()
        if not creator:
            # 如果没有超级用户，创建一个默认用户
            creator = User.objects.create_user(
                username='admin',
                email='admin@example.com',
                password='admin123'
            )
            creator.is_superuser = True
            creator.is_staff = True
            creator.save()
        
        # 保存考核周期
        cycle = serializer.save(created_by=creator)
        
        # 处理考核指标关联
        if 'evaluation_indicators' in serializer.validated_data:
            indicators = serializer.validated_data.pop('evaluation_indicators')
            cycle.evaluation_indicators.set(indicators)
    
    def perform_update(self, serializer):
        """更新考核周期时处理考核指标关联"""
        cycle = serializer.save()
        
        # 处理考核指标关联
        if 'evaluation_indicators' in serializer.validated_data:
            indicators = serializer.validated_data.pop('evaluation_indicators')
            cycle.evaluation_indicators.set(indicators)
    
    @action(detail=True, methods=['post'])
    def generate_tasks(self, request, pk=None):
        """为指定周期生成考核任务"""
        cycle = self.get_object()
        
        # 使用新的任务生成器
        from .services import EvaluationTaskGenerator
        
        generator = EvaluationTaskGenerator(cycle)
        result = generator.generate_tasks()
        
        if result['success']:
            return Response({
                'message': result['message'],
                'tasks_created': result['tasks_created']
            })
        else:
            return Response({
                'error': result['error']
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EvaluationIndicatorViewSet(viewsets.ModelViewSet):
    """考核指标管理"""
    queryset = EvaluationIndicator.objects.all().order_by('category', 'name')
    serializer_class = EvaluationIndicatorSerializer
    # 与前端对齐：类别/启用状态、名称/描述检索
    filterset_fields = ['category', 'is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['category', 'name', 'weight', 'updated_at']


class EvaluationRuleViewSet(viewsets.ModelViewSet):
    """考核规则管理"""
    queryset = EvaluationRule.objects.all().order_by('-created_at')
    serializer_class = EvaluationRuleSerializer
    # 规则为全局配置：支持启用状态/范围筛选，名称与描述检索
    filterset_fields = ['is_active', 'evaluation_scope']
    search_fields = ['name', 'description']
    ordering_fields = ['created_at', 'updated_at', 'name']
    
    @action(detail=False, methods=['post'])
    def create_defaults(self, request):
        """创建默认考核规则"""
        from .services import create_default_evaluation_rules
        
        try:
            create_default_evaluation_rules()
            return Response({
                'message': '默认考核规则创建成功'
            })
        except Exception as e:
            return Response({
                'error': f'创建默认规则失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class ManualEvaluationAssignmentViewSet(viewsets.ModelViewSet):
    """手动评价分配管理"""
    queryset = ManualEvaluationAssignment.objects.all().order_by('-created_at')
    serializer_class = ManualEvaluationAssignmentSerializer
    # 服务端筛选/检索/排序
    filterset_fields = ['cycle', 'evaluator', 'evaluatee', 'relation_type']
    search_fields = ['evaluator__name', 'evaluatee__name', 'reason']
    ordering_fields = ['created_at', 'weight']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        cycle_id = self.request.query_params.get('cycle_id')
        
        if cycle_id:
            queryset = queryset.filter(cycle_id=cycle_id)
            
        return queryset
    
    def create(self, request, *args, **kwargs):
        """创建手动分配"""
        try:
            return super().create(request, *args, **kwargs)
        except Exception as e:
            print(f"创建手动分配错误: {str(e)}")
            return Response({
                'error': f'创建失败: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)

    def perform_create(self, serializer):
        """创建手动分配时自动设置创建者"""
        from django.contrib.auth.models import User
        creator = User.objects.filter(is_superuser=True).first()
        if not creator:
            creator = User.objects.create_user(
                username='admin',
                email='admin@example.com',
                password='admin123'
            )
            creator.is_superuser = True
            creator.is_staff = True
            creator.save()
        
        serializer.save(created_by=creator)
    
    @action(detail=False, methods=['post'], url_path='generate-tasks')
    def generate_tasks(self, request):
        """根据手动分配生成考核任务"""
        try:
            cycle_id = request.data.get('cycle_id')
            if not cycle_id:
                return Response({'error': '请指定考核周期'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 获取该周期的所有手动分配
            assignments = ManualEvaluationAssignment.objects.filter(cycle_id=cycle_id)
            if not assignments.exists():
                return Response({'error': '该周期没有手动分配'}, status=status.HTTP_404_NOT_FOUND)
            
            created_tasks = []
            for assignment in assignments:
                # 检查是否已经存在相同的任务
                existing_task = EvaluationTask.objects.filter(
                    cycle=assignment.cycle,
                    evaluator=assignment.evaluator,
                    evaluatee=assignment.evaluatee,
                    relation_type=assignment.relation_type
                ).first()
                
                if existing_task:
                    continue  # 跳过已存在的任务
                
                # 生成考核码
                evaluation_code = self._generate_evaluation_code()
                
                # 创建考核任务
                task = EvaluationTask.objects.create(
                    cycle=assignment.cycle,
                    evaluator=assignment.evaluator,
                    evaluatee=assignment.evaluatee,
                    relation_type=assignment.relation_type,
                    weight=assignment.weight,
                    evaluation_code=evaluation_code,
                    status='pending'
                )
                created_tasks.append(task)
            
            return Response({
                'message': f'成功生成 {len(created_tasks)} 个考核任务',
                'tasks_count': len(created_tasks)
            })
            
        except Exception as e:
            return Response({
                'error': f'生成任务失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @staticmethod
    def _generate_evaluation_code():
        """生成16位唯一考核码"""
        import string
        import random
        
        characters = string.ascii_uppercase + string.digits
        while True:
            code = ''.join(random.choices(characters, k=16))
            if not EvaluationTask.objects.filter(evaluation_code=code).exists():
                return code


class EvaluationTaskViewSet(viewsets.ModelViewSet):
    """考核任务管理"""
    queryset = EvaluationTask.objects.all().order_by('-assigned_at')
    serializer_class = EvaluationTaskSerializer
    # 基本筛选：周期、评价人、被评价人、关系/状态
    filterset_fields = ['cycle', 'evaluator', 'evaluatee', 'relation_type', 'status']
    search_fields = ['evaluation_code', 'evaluator__name', 'evaluatee__name']
    ordering_fields = ['assigned_at', 'completed_at', 'status']
    
    def create(self, request, *args, **kwargs):
        """创建考核任务时验证考核周期状态"""
        try:
            # 获取考核周期信息
            cycle_id = request.data.get('cycle')
            if cycle_id:
                try:
                    cycle = EvaluationCycle.objects.get(id=cycle_id)
                    if cycle.status != 'active':
                        return Response({
                            'error': '考核周期未激活，无法创建考核任务'
                        }, status=status.HTTP_400_BAD_REQUEST)
                except EvaluationCycle.DoesNotExist:
                    return Response({
                        'error': '考核周期不存在'
                    }, status=status.HTTP_404_NOT_FOUND)
            
            return super().create(request, *args, **kwargs)
        except Exception as e:
            return Response({
                'error': f'创建考核任务失败: {str(e)}'
            }, status=status.HTTP_400_BAD_REQUEST)
    
    def get_queryset(self):
        queryset = super().get_queryset()
        cycle_id = self.request.query_params.get('cycle_id')
        evaluator_id = self.request.query_params.get('evaluator_id')
        evaluatee_id = self.request.query_params.get('evaluatee_id')
        
        if cycle_id:
            queryset = queryset.filter(cycle_id=cycle_id)
        if evaluator_id:
            queryset = queryset.filter(evaluator_id=evaluator_id)
        if evaluatee_id:
            queryset = queryset.filter(evaluatee_id=evaluatee_id)
            
        return queryset
    
    @action(detail=False, methods=['get'], url_path='by-code/(?P<evaluation_code>[^/.]+)')
    def get_by_code(self, request, evaluation_code=None):
        """根据考核码获取任务和相关信息"""
        try:
            # 获取该考核码对应的所有任务
            tasks = EvaluationTask.objects.filter(evaluation_code=evaluation_code)
            if not tasks.exists():
                return Response({'error': '未找到对应的考核任务'}, status=status.HTTP_404_NOT_FOUND)
            
            # 返回第一个任务（通常同一个考核码对应同一个考核人的多个任务）
            task = tasks.first()
            indicators = EvaluationIndicator.objects.filter(is_active=True)
            
            # 检查是否已经有评分记录
            existing_scores = EvaluationScore.objects.filter(task=task)
            scores_data = {score.indicator_id: score for score in existing_scores}
            
            response_data = {
                'task': EvaluationTaskSerializer(task).data,
                'indicators': EvaluationIndicatorSerializer(indicators, many=True).data,
                'existing_scores': {str(k): {
                    'score': v.score,
                    'comment': v.comment
                } for k, v in scores_data.items()}
            }
            
            return Response(response_data)
            
        except EvaluationTask.DoesNotExist:
            return Response({
                'error': '考核任务不存在'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'], url_path='by-task-id/(?P<task_id>[^/.]+)')
    def get_by_task_id(self, request, task_id=None):
        """根据任务ID获取任务和相关信息"""
        try:
            task = EvaluationTask.objects.get(id=task_id)
            indicators = EvaluationIndicator.objects.filter(is_active=True)
            
            # 检查是否已经有评分记录
            existing_scores = EvaluationScore.objects.filter(task=task)
            scores_data = {score.indicator_id: score for score in existing_scores}
            
            response_data = {
                'task': EvaluationTaskSerializer(task).data,
                'indicators': EvaluationIndicatorSerializer(indicators, many=True).data,
                'existing_scores': {str(k): {
                    'score': v.score,
                    'comment': v.comment
                } for k, v in scores_data.items()}
            }
            
            return Response(response_data)
            
        except EvaluationTask.DoesNotExist:
            return Response({
                'error': '考核任务不存在'
            }, status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['get'], url_path='evaluator-tasks/(?P<evaluation_code>[^/.]+)')
    def get_evaluator_tasks(self, request, evaluation_code=None):
        """根据考核码获取考核人的所有待评价任务"""
        try:
            # 查找使用该考核码的所有任务
            tasks = EvaluationTask.objects.filter(evaluation_code=evaluation_code)
            
            if not tasks.exists():
                return Response({
                    'error': '考核码不存在或没有对应的考核任务'
                }, status=status.HTTP_404_NOT_FOUND)
            
            # 获取考核人信息
            evaluator = tasks.first().evaluator
            indicators = EvaluationIndicator.objects.filter(is_active=True)
            
            # 获取所有任务的评分记录
            task_ids = tasks.values_list('id', flat=True)
            existing_scores = EvaluationScore.objects.filter(task_id__in=task_ids)
            scores_data = {}
            for score in existing_scores:
                if score.task_id not in scores_data:
                    scores_data[score.task_id] = {}
                scores_data[score.task_id][score.indicator_id] = {
                    'score': score.score,
                    'comment': score.comment
                }
            
            response_data = {
                'evaluator': {
                    'id': evaluator.id,
                    'name': evaluator.name,
                    'position': evaluator.position_name,
                    'department': evaluator.department_name
                },
                'tasks': EvaluationTaskSerializer(tasks, many=True).data,
                'indicators': EvaluationIndicatorSerializer(indicators, many=True).data,
                'existing_scores': scores_data
            }
            
            return Response(response_data)
            
        except Exception as e:
            return Response({
                'error': f'获取考核任务失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def evaluation_form(self, request, pk=None):
        """获取考核表单"""
        task = self.get_object()
        indicators = EvaluationIndicator.objects.filter(is_active=True)
        
        form_data = {
            'task': EvaluationTaskSerializer(task).data,
            'indicators': EvaluationIndicatorSerializer(indicators, many=True).data
        }
        
        return Response(form_data)
    
    @action(detail=False, methods=['get'], url_path='export-evaluator-codes')
    def export_evaluator_codes(self, request):
        """导出考核人考核码Excel文件"""
        try:
            # 获取考核周期ID
            cycle_id = request.query_params.get('cycle_id')
            if not cycle_id:
                return Response({'error': '请指定考核周期'}, status=status.HTTP_400_BAD_REQUEST)
            
            # 获取该周期的所有任务
            tasks = EvaluationTask.objects.filter(cycle_id=cycle_id).select_related('evaluator')
            
            if not tasks.exists():
                return Response({'error': '该周期没有考核任务'}, status=status.HTTP_404_NOT_FOUND)
            
            # 按考核人分组，每个考核人只取一个考核码
            evaluator_data = {}
            for task in tasks:
                evaluator = task.evaluator
                if evaluator.id not in evaluator_data:
                    evaluator_data[evaluator.id] = {
                        'name': evaluator.name,
                        'position': evaluator.position_name,
                        'department': evaluator.department_name,
                        'evaluation_code': task.evaluation_code
                    }
            
            # 创建Excel文件
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "考核码分发表"
            
            # 设置标题
            ws['A1'] = '考核码分发表'
            ws['A1'].font = Font(size=20, bold=True)
            ws['A1'].alignment = Alignment(horizontal='center')
            ws.merge_cells('A1:E1')
            
            # 设置表头
            headers = ['序号', '考核人', '职位', '部门', '考核码']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(size=14, bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            
            # 填充数据
            row = 4
            for idx, (evaluator_id, data) in enumerate(evaluator_data.items(), 1):
                ws.cell(row=row, column=1, value=idx).font = Font(size=12)
                ws.cell(row=row, column=2, value=data['name']).font = Font(size=12)
                ws.cell(row=row, column=3, value=data['position']).font = Font(size=12)
                ws.cell(row=row, column=4, value=data['department']).font = Font(size=12)
                ws.cell(row=row, column=5, value=data['evaluation_code']).font = Font(size=12, bold=True)
                row += 1
            
            # 设置列宽
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            
            # 设置行高
            for row_num in range(1, row):
                ws.row_dimensions[row_num].height = 25
            
            # 创建HTTP响应
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="考核码分发表_{cycle_id}.xlsx"'
            
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({
                'error': f'导出失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EvaluationScoreViewSet(viewsets.ModelViewSet):
    """考核评分管理"""
    queryset = EvaluationScore.objects.all()
    serializer_class = EvaluationScoreSerializer
    filterset_fields = ['task', 'indicator']
    search_fields = ['comment', 'task__evaluation_code', 'indicator__name']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        task_id = self.request.query_params.get('task_id')
        
        if task_id:
            queryset = queryset.filter(task_id=task_id)
            
        return queryset
    
    @action(detail=False, methods=['post'])
    def submit_task_scores(self, request):
        """提交任务的所有评分"""
        task_id = request.data.get('task_id')
        scores_data = request.data.get('scores', [])
        
        try:
            task = EvaluationTask.objects.get(id=task_id)
            
            # 获取评价人的职级权重
            evaluator_weight = 1.0
            try:
                weight_config = PositionWeight.objects.filter(
                    position_id=task.evaluator.position_id,
                    is_active=True
                ).first()
                if weight_config:
                    evaluator_weight = float(weight_config.weight)
            except:
                pass
            
            # 清除原有评分
            EvaluationScore.objects.filter(task=task).delete()
            
            # 创建新评分
            created_scores = []
            for score_data in scores_data:
                # 计算加权评分
                weighted_score = float(score_data['score']) * evaluator_weight
                
                score = EvaluationScore.objects.create(
                    task=task,
                    indicator_id=score_data['indicator_id'],
                    score=score_data['score'],
                    weighted_score=weighted_score,
                    comment=score_data.get('comment', '')
                )
                created_scores.append(score)
            
            # 更新任务状态为已完成
            task.status = 'completed'
            from django.utils import timezone
            task.completed_at = timezone.now()
            task.save()
            
            return Response({
                'message': '评分提交成功',
                'scores_count': len(created_scores),
                'evaluator_weight': evaluator_weight
            })
            
        except EvaluationTask.DoesNotExist:
            return Response({
                'error': '考核任务不存在'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'error': f'提交评分失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class EvaluationResultViewSet(viewsets.ModelViewSet):
    """考核结果管理"""
    queryset = EvaluationResult.objects.all().order_by('-calculated_at')
    serializer_class = EvaluationResultSerializer
    # 支持周期与被评价人筛选
    filterset_fields = ['cycle', 'employee']
    search_fields = ['employee__name']
    ordering_fields = ['weighted_score', 'total_score', 'rank', 'calculated_at']
    
    def get_queryset(self):
        queryset = super().get_queryset()
        cycle_id = self.request.query_params.get('cycle_id')
        evaluatee_id = self.request.query_params.get('evaluatee_id')
        
        if cycle_id:
            queryset = queryset.filter(cycle_id=cycle_id)
        if evaluatee_id:
            queryset = queryset.filter(employee_id=evaluatee_id)
            
        return queryset
    
    @action(detail=False, methods=['post'])
    def calculate_cycle_results(self, request):
        """计算指定周期的考核结果"""
        cycle_id = request.data.get('cycle_id')
        
        if not cycle_id:
            return Response({
                'error': '缺少周期 ID'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            cycle = EvaluationCycle.objects.get(id=cycle_id)
            
            # 获取该周期下所有已完成的任务
            completed_tasks = EvaluationTask.objects.filter(
                cycle=cycle, 
                status='completed'
            ).select_related('evaluatee', 'evaluator')
            
            if not completed_tasks.exists():
                return Response({
                    'message': '该周期没有已完成的考核任务'
                })
            
            # 按被评价人分组计算结果
            from django.db.models import Avg, Count
            from decimal import Decimal
            
            # 先清除旧的结果
            EvaluationResult.objects.filter(cycle=cycle).delete()
            
            results_created = 0
            
            # 获取所有被评价人
            evaluatees = set(task.evaluatee for task in completed_tasks)
            
            for evaluatee in evaluatees:
                # 获取该员工的所有评分
                employee_tasks = completed_tasks.filter(evaluatee=evaluatee)
                
                if not employee_tasks.exists():
                    continue
                
                # 计算加权平均分
                total_weighted_score = Decimal('0')
                total_weight = Decimal('0')
                
                relation_scores = {}
                
                for task in employee_tasks:
                    # 获取任务的所有评分
                    task_scores = EvaluationScore.objects.filter(task=task)
                    
                    if task_scores.exists():
                        # 计算任务的加权平均分
                        task_total_score = Decimal('0')
                        task_total_weight = Decimal('0')
                        
                        for score in task_scores:
                            indicator_weight = Decimal(str(score.indicator.weight))
                            weighted_score = Decimal(str(score.score)) * indicator_weight
                            task_total_score += weighted_score
                            task_total_weight += indicator_weight
                        
                        if task_total_weight > 0:
                            task_avg_score = task_total_score / task_total_weight
                            task_weight = Decimal(str(task.weight))
                            
                            total_weighted_score += task_avg_score * task_weight
                            total_weight += task_weight
                            
                            # 按关系类型分组记录
                            relation_type = task.relation_type
                            if relation_type not in relation_scores:
                                relation_scores[relation_type] = []
                            relation_scores[relation_type].append(task_avg_score)
                
                if total_weight > 0:
                    weighted_score = total_weighted_score / total_weight
                    
                    # 计算各关系类型的平均分
                    superior_score = None
                    peer_score = None
                    subordinate_score = None
                    
                    if 'superior' in relation_scores:
                        superior_score = sum(relation_scores['superior']) / len(relation_scores['superior'])
                    if 'peer' in relation_scores:
                        peer_score = sum(relation_scores['peer']) / len(relation_scores['peer'])
                    if 'subordinate' in relation_scores:
                        subordinate_score = sum(relation_scores['subordinate']) / len(relation_scores['subordinate'])
                    
                    # 创建结果记录
                    EvaluationResult.objects.create(
                        cycle=cycle,
                        employee=evaluatee,
                        total_score=total_weighted_score,
                        weighted_score=weighted_score,
                        superior_score=superior_score,
                        peer_score=peer_score,
                        subordinate_score=subordinate_score,
                        is_final=True
                    )
                    results_created += 1
            
            # 计算排名
            results = EvaluationResult.objects.filter(cycle=cycle).order_by('-weighted_score')
            for index, result in enumerate(results, 1):
                result.rank = index
                result.save()
            
            return Response({
                'message': f'成功计算 {results_created} 个员工的考核结果',
                'results_created': results_created
            })
            
        except EvaluationCycle.DoesNotExist:
            return Response({
                'error': '考核周期不存在'
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({
                'error': f'计算结果失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def export(self, request):
        """导出考核结果为 Excel (xlsx)。支持 ?cycle= 或 ?cycle_id= 过滤。"""
        try:
            cycle_id = request.query_params.get('cycle') or request.query_params.get('cycle_id')
            qs = self.get_queryset()
            if cycle_id:
                qs = qs.filter(cycle_id=cycle_id)

            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '考核结果'
            headers = ['周期ID', '周期名称', '员工ID', '员工姓名', '加权平均分', '总分', '上级评分', '同级评分', '下级评分', '排名', '是否最终', '计算时间']
            ws.append(headers)

            # 批量预取
            qs = qs.select_related('cycle', 'employee')
            for r in qs:
                ws.append([
                    r.cycle_id,
                    getattr(r.cycle, 'name', ''),
                    r.employee_id,
                    getattr(r.employee, 'name', ''),
                    float(r.weighted_score or 0),
                    float(r.total_score or 0),
                    float(r.superior_score or 0) if r.superior_score is not None else None,
                    float(r.peer_score or 0) if r.peer_score is not None else None,
                    float(r.subordinate_score or 0) if r.subordinate_score is not None else None,
                    r.rank,
                    '是' if r.is_final else '否',
                    r.calculated_at.strftime('%Y-%m-%d %H:%M:%S') if r.calculated_at else ''
                ])

            from io import BytesIO
            bio = BytesIO()
            wb.save(bio)
            bio.seek(0)

            filename = 'results.xlsx' if not cycle_id else f'results_cycle_{cycle_id}.xlsx'
            resp = HttpResponse(bio.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            resp['Content-Disposition'] = f'attachment; filename="{filename}"'
            return resp
        except Exception as e:
            return Response({'error': f'导出失败: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def overview_stats(request):
    """获取系统概览统计信息"""
    from django.db.models import Avg, Count
    from decimal import Decimal
    
    # 基础统计
    total_cycles = EvaluationCycle.objects.count()
    active_cycles = EvaluationCycle.objects.filter(status='active').count()
    total_tasks = EvaluationTask.objects.count()
    completed_tasks = EvaluationTask.objects.filter(status='completed').count()
    pending_tasks = EvaluationTask.objects.filter(status='pending').count()
    
    # 计算完成率
    completion_rate = 0
    if total_tasks > 0:
        completion_rate = (completed_tasks / total_tasks) * 100
    
    # 计算平均分数
    average_score = 0
    results = EvaluationResult.objects.all()
    if results.exists():
        total_score = sum(float(result.weighted_score) for result in results)
        average_score = total_score / results.count()
    
    # 获取员工总数
    total_employees = Employee.objects.filter(is_active=True).count()
    if total_employees == 0:
        # 如果本地没有员工，尝试从组织架构中台获取
        try:
            org_api_url = getattr(settings, 'ORG_PLATFORM_API_URL', 'http://localhost:8000/api')
            response = requests.get(f'{org_api_url}/employees/')
            if response.status_code == 200:
                employees_data = response.json()
                total_employees = employees_data.get('count', 0)
        except:
            pass
    
    # 计算部门统计
    department_stats = []
    departments = Employee.objects.values('department_name').annotate(
        employee_count=Count('id')
    ).filter(department_name__isnull=False, is_active=True)
    
    for dept in departments:
        dept_name = dept['department_name']
        employee_count = dept['employee_count']
        
        # 部门任务统计
        dept_employees = Employee.objects.filter(department_name=dept_name, is_active=True)
        dept_employee_ids = list(dept_employees.values_list('id', flat=True))
        dept_total_tasks = EvaluationTask.objects.filter(evaluatee_id__in=dept_employee_ids).count()
        dept_completed_tasks = EvaluationTask.objects.filter(
            evaluatee_id__in=dept_employee_ids, 
            status='completed'
        ).count()
        
        dept_completion_rate = 0
        if dept_total_tasks > 0:
            dept_completion_rate = (dept_completed_tasks / dept_total_tasks) * 100
        
        # 部门平均分数
        dept_results = EvaluationResult.objects.filter(employee_id__in=dept_employee_ids)
        dept_average_score = 0
        if dept_results.exists():
            dept_total_score = sum(float(result.weighted_score) for result in dept_results)
            dept_average_score = dept_total_score / dept_results.count()
        
        department_stats.append({
            'department_name': dept_name,
            'employee_count': employee_count,
            'completion_rate': round(dept_completion_rate, 1),
            'average_score': round(dept_average_score, 1)
        })
    
    # 按平均分数排序
    department_stats.sort(key=lambda x: x['average_score'], reverse=True)
    
    stats = {
        'total_cycles': total_cycles,
        'active_cycles': active_cycles,
        'total_tasks': total_tasks,
        'completed_tasks': completed_tasks,
        'total_employees': total_employees,
        'pending_evaluations': pending_tasks,
        'completion_rate': round(completion_rate, 1),
        'average_score': round(average_score, 1),
        'department_stats': department_stats
    }
    
    return Response(stats)


@api_view(['GET'])
def cycle_stats(request, cycle_id):
    """获取指定周期的统计信息"""
    try:
        cycle = EvaluationCycle.objects.get(id=cycle_id)
        
        stats = {
            'cycle': EvaluationCycleSerializer(cycle).data,
            'total_tasks': EvaluationTask.objects.filter(cycle=cycle).count(),
            'completed_tasks': EvaluationTask.objects.filter(cycle=cycle, status='completed').count(),
            'pending_tasks': EvaluationTask.objects.filter(cycle=cycle, status='pending').count(),
            'total_results': EvaluationResult.objects.filter(cycle=cycle).count(),
        }
        
        return Response(stats)
        
    except EvaluationCycle.DoesNotExist:
        return Response({
            'error': '考核周期不存在'
        }, status=status.HTTP_404_NOT_FOUND)


class EmployeeViewSet(viewsets.ModelViewSet):
    """员工管理（本地副本）"""
    queryset = Employee.objects.all().order_by('id')
    serializer_class = EmployeeSerializer
    # 服务端筛选/检索/排序
    filterset_fields = ['department_name', 'status', 'is_active']
    search_fields = ['name', 'email', 'phone', 'employee_id']
    ordering_fields = ['department_name', 'position_level', 'employee_id', 'name']
    
    @action(detail=False, methods=['post'])
    def sync(self, request):
        """同步组织架构中台的员工数据"""
        try:
            org_api_url = getattr(settings, 'ORG_PLATFORM_API_URL', 'http://localhost:8000/api')
            response = requests.get(f'{org_api_url}/employees/')
            
            if response.status_code == 200:
                employees_data = response.json().get('results', [])
                
                synced_count = 0
                for emp_data in employees_data:
                    employee, created = Employee.objects.get_or_create(
                        employee_id=emp_data['employee_id'],
                        defaults={
                            'name': emp_data['name'],
                            'department_id': emp_data['department'],
                            'department_name': emp_data.get('department_name', ''),
                            'position_id': emp_data['position'],
                            'position_name': emp_data.get('position_name', ''),
                            'status': emp_data.get('status', 'active')
                        }
                    )
                    
                    if not created:
                        # 更新现有员工信息
                        employee.name = emp_data['name']
                        employee.department_id = emp_data['department']
                        employee.department_name = emp_data.get('department_name', '')
                        employee.position_id = emp_data['position']
                        employee.position_name = emp_data.get('position_name', '')
                        employee.status = emp_data.get('status', 'active')
                        employee.save()
                    
                    synced_count += 1
                
                return Response({
                    'message': f'成功同步 {synced_count} 个员工数据',
                    'synced_count': synced_count
                })
            else:
                return Response({
                    'error': '无法连接到组织架构中台'
                }, status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({
                'error': f'同步员工数据失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class PositionWeightViewSet(viewsets.ModelViewSet):
    """职级权重管理"""
    queryset = PositionWeight.objects.all().order_by('-created_at')
    serializer_class = PositionWeightSerializer
    filterset_fields = ['is_active']
    search_fields = ['position__name']
    
    @action(detail=False, methods=['post'], url_path='bulk-update')
    def bulk_update_weights(self, request):
        """批量更新职级权重"""
        try:
            weight_data = request.data.get('weights', [])
            if not weight_data:
                return Response({'error': '请提供权重数据'}, status=status.HTTP_400_BAD_REQUEST)
            
            updated_count = 0
            for item in weight_data:
                position_id = item.get('position_id')
                position_name = item.get('position_name', '')
                position_level = item.get('position_level', 0)
                weight = item.get('weight')
                is_active = item.get('is_active', True)
                
                if not position_id or not weight:
                    continue
                
                weight_obj, created = PositionWeight.objects.get_or_create(
                    position_id=position_id,
                    defaults={
                        'position_name': position_name,
                        'position_level': position_level,
                        'weight': weight, 
                        'is_active': is_active
                    }
                )
                
                if not created:
                    weight_obj.position_name = position_name
                    weight_obj.position_level = position_level
                    weight_obj.weight = weight
                    weight_obj.is_active = is_active
                    weight_obj.save()
                    updated_count += 1
                else:
                    updated_count += 1
            
            return Response({
                'message': f'成功更新 {updated_count} 个职级权重',
                'updated_count': updated_count
            })
            
        except Exception as e:
            return Response({
                'error': f'批量更新失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'], url_path='default-weights')
    def get_default_weights(self, request):
        """获取默认权重配置建议"""
        try:
            # 通过API获取组织架构系统中的职位信息
            import requests
            from django.conf import settings
            
            positions = []
            try:
                # 从组织架构系统获取职位列表
                url = f"{settings.ORG_PLATFORM_URL}/api/positions/"
                print(f"正在请求职位数据: {url}")
                response = requests.get(url, timeout=10)
                print(f"API响应状态: {response.status_code}")
                
                if response.status_code == 200:
                    data = response.json()
                    positions = data.get('results', [])
                    print(f"获取到 {len(positions)} 个职位")
                else:
                    print(f"API请求失败: {response.status_code}")
                    positions = []
            except Exception as e:
                print(f"API请求异常: {str(e)}")
                positions = []
            
            # 如果没有从API获取到数据，使用模拟数据
            if not positions:
                print("使用模拟职位数据")
                positions = [
                    {'id': 1, 'name': '董事长', 'level': 13},
                    {'id': 2, 'name': '总经理', 'level': 12},
                    {'id': 3, 'name': '副总经理', 'level': 11},
                    {'id': 4, 'name': '部门经理', 'level': 9},
                    {'id': 5, 'name': '部门副经理', 'level': 8},
                    {'id': 6, 'name': '主管', 'level': 4},
                    {'id': 7, 'name': '专员', 'level': 2},
                    {'id': 8, 'name': '助理', 'level': 1},
                ]
            
            default_weights = []
            
            for position in positions:
                # 根据职位级别设置默认权重
                level = position.get('level', 0)
                if level >= 13:  # 高层正职
                    weight = 1.8
                elif level >= 11:  # 高层副职
                    weight = 1.6
                elif level >= 9:  # 中层正职
                    weight = 1.2
                elif level >= 8:  # 中层副职
                    weight = 1.1
                elif level >= 4:  # 基层正职
                    weight = 1.0
                else:  # 普通员工
                    weight = 0.9
                
                # 检查是否已有权重配置
                current_weight = None
                try:
                    existing_weight = PositionWeight.objects.filter(
                        position_id=position.get('id')
                    ).first()
                    if existing_weight:
                        current_weight = float(existing_weight.weight)
                except:
                    pass
                
                default_weights.append({
                    'position_id': position.get('id'),
                    'position_name': position.get('name'),
                    'position_level': level,
                    'suggested_weight': weight,
                    'current_weight': current_weight
                })
            
            print(f"返回 {len(default_weights)} 个默认权重配置")
            return Response({
                'default_weights': default_weights
            })
            
        except Exception as e:
            print(f"获取默认权重异常: {str(e)}")
            return Response({
                'error': f'获取默认权重失败: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def stats_overview(request):
    """获取系统概览统计数据"""
    try:
        # 基础统计
        total_cycles = EvaluationCycle.objects.count()
        active_cycles = EvaluationCycle.objects.filter(status='active').count()
        completed_cycles = EvaluationCycle.objects.filter(status='completed').count()
        
        # 任务统计
        total_tasks = EvaluationTask.objects.count()
        completed_tasks = EvaluationTask.objects.filter(status='completed').count()
        pending_tasks = EvaluationTask.objects.filter(status='pending').count()
        
        # 计算完成率
        completion_rate = 0
        if total_tasks > 0:
            completion_rate = round((completed_tasks / total_tasks) * 100, 1)
        
        # 评分统计
        avg_score = 0
        avg_grade = 'B'
        score_distribution = {'A+': 0, 'A': 0, 'A-': 0, 'B+': 0, 'B': 0, 'B-': 0, 'C': 0}
        
        if completed_tasks > 0:
            # 计算平均加权评分
            score_stats = EvaluationScore.objects.filter(
                task__status='completed'
            ).aggregate(
                avg_weighted_score=Avg('weighted_score'),
                avg_score=Avg('score')
            )
            avg_score = round(score_stats['avg_score'] or 0, 1)
            
            # 根据平均分计算等级
            if avg_score >= 90:
                avg_grade = 'A+'
            elif avg_score >= 85:
                avg_grade = 'A'
            elif avg_score >= 80:
                avg_grade = 'A-'
            elif avg_score >= 75:
                avg_grade = 'B+'
            elif avg_score >= 70:
                avg_grade = 'B'
            elif avg_score >= 65:
                avg_grade = 'B-'
            else:
                avg_grade = 'C'
            
            # 计算评分分布
            tasks_with_scores = EvaluationTask.objects.filter(
                status='completed',
                evaluationscore__isnull=False
            ).distinct()
            
            for task in tasks_with_scores:
                task_avg_score = task.evaluationscore_set.aggregate(avg=Avg('score'))['avg'] or 0
                if task_avg_score >= 90:
                    score_distribution['A+'] += 1
                elif task_avg_score >= 85:
                    score_distribution['A'] += 1
                elif task_avg_score >= 80:
                    score_distribution['A-'] += 1
                elif task_avg_score >= 75:
                    score_distribution['B+'] += 1
                elif task_avg_score >= 70:
                    score_distribution['B'] += 1
                elif task_avg_score >= 65:
                    score_distribution['B-'] += 1
                else:
                    score_distribution['C'] += 1
        
        # 部门绩效统计
        dept_performance = []
        dept_stats = {}
        
        # 按部门统计任务和评分
        for task in EvaluationTask.objects.filter(status='completed'):
            dept = task.evaluatee.department_name or '未分配'
            if dept not in dept_stats:
                dept_stats[dept] = {
                    'total_tasks': 0,
                    'completed_tasks': 0,
                    'total_score': 0,
                    'avg_score': 0
                }
            
            dept_stats[dept]['total_tasks'] += 1
            dept_stats[dept]['completed_tasks'] += 1
            
            # 计算部门平均分
            task_scores = task.evaluationscore_set.aggregate(avg=Avg('score'))['avg'] or 0
            dept_stats[dept]['total_score'] += task_scores
        
        # 计算各部门平均分
        for dept, stats in dept_stats.items():
            if stats['completed_tasks'] > 0:
                stats['avg_score'] = round(stats['total_score'] / stats['completed_tasks'], 1)
                dept_performance.append({
                    'department': dept,
                    'total_tasks': stats['total_tasks'],
                    'completed_tasks': stats['completed_tasks'],
                    'completion_rate': round((stats['completed_tasks'] / stats['total_tasks']) * 100, 1),
                    'avg_score': stats['avg_score']
                })
        
        # 异常检测（评分波动较大的员工）
        anomaly_count = 0
        anomaly_rate = 0
        if completed_tasks > 0:
            # 计算评分标准差，识别异常
            scores = list(EvaluationScore.objects.filter(
                task__status='completed'
            ).values_list('score', flat=True))
            
            if len(scores) > 1:
                import statistics
                mean_score = statistics.mean(scores)
                std_score = statistics.stdev(scores)
                
                # 识别超出2个标准差的异常评分
                for score in scores:
                    if abs(score - mean_score) > 2 * std_score:
                        anomaly_count += 1
                
                anomaly_rate = round((anomaly_count / len(scores)) * 100, 1)
        
        # 最近活动
        recent_activities = []
        recent_tasks = EvaluationTask.objects.filter(
            status='completed'
        ).order_by('-completed_at')[:5]
        
        for task in recent_tasks:
            if task.completed_at:
                time_diff = timezone.now() - task.completed_at
                if time_diff.days > 0:
                    time_text = f"{time_diff.days}天前"
                elif time_diff.seconds > 3600:
                    time_text = f"{time_diff.seconds // 3600}小时前"
                else:
                    time_text = f"{time_diff.seconds // 60}分钟前"
                
                recent_activities.append({
                    'id': task.id,
                    'type': 'task_completed',
                    'message': f"{task.evaluator.name}完成了对{task.evaluatee.name}的考核",
                    'time': time_text,
                    'timestamp': task.completed_at.isoformat()
                })
        
        # 考核周期进度
        cycle_progress = []
        for cycle in EvaluationCycle.objects.filter(status__in=['active', 'completed'])[:5]:
            cycle_tasks = EvaluationTask.objects.filter(cycle=cycle)
            cycle_completed = cycle_tasks.filter(status='completed').count()
            cycle_total = cycle_tasks.count()
            progress = 0
            if cycle_total > 0:
                progress = round((cycle_completed / cycle_total) * 100)
            
            cycle_progress.append({
                'id': cycle.id,
                'name': cycle.name,
                'status': cycle.status,
                'progress': progress,
                'start_date': cycle.start_date.isoformat() if cycle.start_date else None,
                'end_date': cycle.end_date.isoformat() if cycle.end_date else None
            })
        
        # 绩效趋势（最近7天的完成情况）
        performance_trend = []
        for i in range(7):
            date = timezone.now().date() - timedelta(days=i)
            daily_completed = EvaluationTask.objects.filter(
                status='completed',
                completed_at__date=date
            ).count()
            performance_trend.append({
                'date': date.isoformat(),
                'completed': daily_completed
            })
        performance_trend.reverse()  # 按时间正序排列
        
        return Response({
            'total_cycles': total_cycles,
            'active_cycles': active_cycles,
            'completed_cycles': completed_cycles,
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'completion_rate': completion_rate,
            'avg_score': avg_score,
            'avg_grade': avg_grade,
            'score_distribution': score_distribution,
            'dept_performance': dept_performance,
            'anomaly_count': anomaly_count,
            'anomaly_rate': anomaly_rate,
            'recent_activities': recent_activities,
            'cycle_progress': cycle_progress,
            'performance_trend': performance_trend
        })
        
    except Exception as e:
        return Response({
            'error': f'获取统计数据失败: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['GET'])
def stats_cycle(request, cycle_id):
    """获取特定考核周期的统计数据"""
    try:
        cycle = EvaluationCycle.objects.get(id=cycle_id)
        tasks = EvaluationTask.objects.filter(cycle=cycle)
        
        # 基础统计
        total_tasks = tasks.count()
        completed_tasks = tasks.filter(status='completed').count()
        pending_tasks = tasks.filter(status='pending').count()
        
        # 完成率
        completion_rate = 0
        if total_tasks > 0:
            completion_rate = round((completed_tasks / total_tasks) * 100, 1)
        
        # 评分统计
        scores = EvaluationScore.objects.filter(task__cycle=cycle)
        avg_score = 0
        avg_grade = 'B'
        
        if scores.exists():
            score_stats = scores.aggregate(
                avg_weighted_score=Avg('weighted_score'),
                avg_score=Avg('score')
            )
            avg_score = round(score_stats['avg_score'] or 0, 1)
            
            # 等级计算
            if avg_score >= 90:
                avg_grade = 'A+'
            elif avg_score >= 85:
                avg_grade = 'A'
            elif avg_score >= 80:
                avg_grade = 'A-'
            elif avg_score >= 75:
                avg_grade = 'B+'
            elif avg_score >= 70:
                avg_grade = 'B'
            elif avg_score >= 65:
                avg_grade = 'B-'
            else:
                avg_grade = 'C'
        
        # 部门统计
        dept_stats = {}
        for task in tasks.filter(status='completed'):
            dept = task.evaluatee.department or '未分配'
            if dept not in dept_stats:
                dept_stats[dept] = {'total': 0, 'completed': 0}
            dept_stats[dept]['total'] += 1
            if task.status == 'completed':
                dept_stats[dept]['completed'] += 1
        
        # 计算部门完成度
        dept_completion = []
        for dept, stats in dept_stats.items():
            completion = 0
            if stats['total'] > 0:
                completion = round((stats['completed'] / stats['total']) * 100, 1)
            dept_completion.append({
                'department': dept,
                'total': stats['total'],
                'completed': stats['completed'],
                'completion_rate': completion
            })
        
        return Response({
            'cycle': {
                'id': cycle.id,
                'name': cycle.name,
                'status': cycle.status,
                'start_date': cycle.start_date.isoformat() if cycle.start_date else None,
                'end_date': cycle.end_date.isoformat() if cycle.end_date else None
            },
            'total_tasks': total_tasks,
            'completed_tasks': completed_tasks,
            'pending_tasks': pending_tasks,
            'completion_rate': completion_rate,
            'avg_score': avg_score,
            'avg_grade': avg_grade,
            'dept_completion': dept_completion
        })
        
    except EvaluationCycle.DoesNotExist:
        return Response({
            'error': '考核周期不存在'
        }, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        return Response({
            'error': f'获取周期统计失败: {str(e)}'
        }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
